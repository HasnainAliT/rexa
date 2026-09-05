"""Parse CSV / Excel class rosters into column + row maps."""
from __future__ import annotations

import csv
import io
from typing import Any


def parse_tabular_bytes(data: bytes, filename: str) -> dict[str, Any]:
    name = (filename or "roster.csv").lower()
    if name.endswith(".xls") and not name.endswith(".xlsx") and not name.endswith(".xlsm"):
        raise ValueError(
            "Save the file as Excel .xlsx (not the older .xls format) and upload again."
        )
    if name.endswith(".xlsx") or name.endswith(".xlsm"):
        return _parse_xlsx(data, filename)
    return _parse_csv(data, filename)


def _filled_count(row: tuple[Any, ...] | None) -> int:
    if not row:
        return 0
    return sum(1 for cell in row if cell is not None and str(cell).strip())


def _normalize_columns(header_row: tuple[Any, ...]) -> list[str]:
    columns: list[str] = []
    seen: dict[str, int] = {}
    for index, cell in enumerate(header_row):
        label = str(cell).strip() if cell is not None else ""
        base = label or f"Column {index + 1}"
        count = seen.get(base, 0)
        seen[base] = count + 1
        columns.append(base if count == 0 else f"{base} ({count + 1})")
    return columns


def _parse_csv(data: bytes, filename: str) -> dict[str, Any]:
    text = data.decode("utf-8-sig", errors="replace")
    sample = text[:4096]
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=",;\t")
    except csv.Error:
        dialect = csv.excel
    reader = csv.reader(io.StringIO(text), dialect=dialect)
    raw_rows = [tuple(row) for row in reader]
    return _rows_from_matrix(raw_rows, filename)


def _parse_xlsx(data: bytes, filename: str) -> dict[str, Any]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:  # pragma: no cover
        raise ValueError("Excel support requires openpyxl.") from exc

    workbook = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    sheet = workbook.active
    if sheet is None:
        workbook.close()
        raise ValueError("This workbook has no worksheets.")

    matrix = [tuple(row) for row in sheet.iter_rows(values_only=True)]
    workbook.close()
    return _rows_from_matrix(matrix, filename)


def _rows_from_matrix(matrix: list[tuple[Any, ...]], filename: str) -> dict[str, Any]:
    header_index = 0
    best = 0
    for index, row in enumerate(matrix[:8]):
        filled = _filled_count(row)
        if filled > best:
            best = filled
            header_index = index
        if filled >= 3:
            break
    if best < 2:
        raise ValueError("No header row found in this file.")

    header_row = matrix[header_index]
    columns = _normalize_columns(header_row)
    rows: list[dict[str, str]] = []
    for values in matrix[header_index + 1 :]:
        row = {
            columns[i]: (
                ""
                if i >= len(values) or values[i] is None
                else str(values[i]).strip()
            )
            for i in range(len(columns))
        }
        if any(row.values()):
            rows.append(row)

    if not rows:
        raise ValueError("No data rows found in this file.")
    return {
        "filename": filename,
        "columns": columns,
        "rows": rows,
        "row_count": len(rows),
    }
