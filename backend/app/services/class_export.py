"""Class-level Excel and PDF exports for batch grading results."""
from __future__ import annotations

import io
from xml.sax.saxutils import escape

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import landscape, letter
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import inch
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle


HEADERS = [
    "Student name",
    "Student ID",
    "Class",
    "Question / assignment",
    "Role coverage %",
    "Concept coverage %",
    "Depth %",
    "Stars",
    "Overall",
    "Status",
]


def generate_class_xlsx(rows: list[dict], title: str = "RExA class report") -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Class report"
    sheet["A1"] = title
    sheet["A1"].font = Font(bold=True, color="1E1B4B", size=14)
    sheet.merge_cells("A1:J1")
    sheet["A2"] = (
        "Pass requires the assignment role-coverage, concept-coverage, and star thresholds."
    )
    sheet.merge_cells("A2:J2")

    header_fill = PatternFill("solid", fgColor="EEF2FF")
    fail_fill = PatternFill("solid", fgColor="FFF7ED")
    header_font = Font(bold=True, color="1E1B4B")

    for col, header in enumerate(HEADERS, start=1):
        cell = sheet.cell(row=4, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(wrap_text=True, vertical="center")

    for index, row in enumerate(rows, start=5):
        values = [
            row.get("student_name") or "",
            row.get("student_id") or "",
            row.get("class_name") or "",
            row.get("question") or "",
            row.get("role_coverage") or 0,
            row.get("concept_coverage") or 0,
            row.get("depth") or 0,
            row.get("stars") or 0,
            row.get("overall") or 0,
            row.get("status") or "",
        ]
        failed = "below" in str(row.get("status") or "").lower()
        for col, value in enumerate(values, start=1):
            cell = sheet.cell(row=index, column=col, value=value)
            if failed:
                cell.fill = fail_fill

    widths = [22, 14, 18, 42, 16, 18, 12, 10, 12, 16]
    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(index)].width = width

    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def generate_class_pdf(rows: list[dict], title: str = "RExA class report") -> bytes:
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(letter),
        leftMargin=0.5 * inch,
        rightMargin=0.5 * inch,
        topMargin=0.55 * inch,
        bottomMargin=0.5 * inch,
        title=title,
    )
    styles = getSampleStyleSheet()
    heading = ParagraphStyle(
        "ClassTitle",
        parent=styles["Heading1"],
        textColor=colors.HexColor("#1e1b4b"),
        fontSize=16,
        spaceAfter=6,
    )
    body = ParagraphStyle(
        "ClassBody",
        parent=styles["Normal"],
        fontSize=8,
        leading=11,
        textColor=colors.HexColor("#1e1b4b"),
    )
    story = [
        Paragraph(escape(title), heading),
        Paragraph(
            "Pass requires the assignment role-coverage, concept-coverage, and star thresholds. "
            "Rows in orange are below threshold.",
            body,
        ),
        Spacer(1, 10),
    ]

    data = [[Paragraph(f"<b>{escape(h)}</b>", body) for h in HEADERS]]
    for row in rows:
        data.append(
            [
                Paragraph(escape(str(row.get("student_name") or "—")), body),
                Paragraph(escape(str(row.get("student_id") or "—")), body),
                Paragraph(escape(str(row.get("class_name") or "—")), body),
                Paragraph(escape(str(row.get("question") or "—")), body),
                Paragraph(f"{row.get('role_coverage') or 0}%", body),
                Paragraph(f"{row.get('concept_coverage') or 0}%", body),
                Paragraph(f"{row.get('depth') or 0}%", body),
                Paragraph(str(row.get("stars") or 0), body),
                Paragraph(str(row.get("overall") or 0), body),
                Paragraph(escape(str(row.get("status") or "")), body),
            ]
        )

    table = Table(
        data,
        colWidths=[1.3 * inch, 0.9 * inch, 1.1 * inch, 2.4 * inch, 0.85 * inch, 0.95 * inch, 0.7 * inch, 0.6 * inch, 0.7 * inch, 1.0 * inch],
        repeatRows=1,
    )
    style_commands = [
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#eef2ff")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.HexColor("#1e1b4b")),
        ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#c7d2fe")),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("LEFTPADDING", (0, 0), (-1, -1), 4),
        ("RIGHTPADDING", (0, 0), (-1, -1), 4),
        ("TOPPADDING", (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
    ]
    for index, row in enumerate(rows, start=1):
        if "below" in str(row.get("status") or "").lower():
            style_commands.append(
                ("BACKGROUND", (0, index), (-1, index), colors.HexColor("#fff7ed"))
            )
    table.setStyle(TableStyle(style_commands))
    story.append(table)
    doc.build(story)
    return buffer.getvalue()
